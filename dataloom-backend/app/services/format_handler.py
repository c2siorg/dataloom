"""Multi-format file ingestion and export for DataLoom.

Supports CSV, XLSX, JSON, Parquet, and TSV formats.
Each loader returns a pandas DataFrame; each exporter serialises a DataFrame to disk.
"""

from pathlib import Path

import pandas as pd

from app.utils.logging import get_logger

logger = get_logger(__name__)


class FormatError(Exception):
    """Raised when a file cannot be parsed in the expected format."""

    pass


_EXTENSION_MAP: dict[str, str] = {
    ".csv": "csv",
    ".xlsx": "xlsx",
    ".json": "json",
    ".parquet": "parquet",
    ".tsv": "tsv",
}


def detect_format(filename: str) -> str:
    """Auto-detect file format from its extension.

    Args:
        filename: Filename or path to inspect.

    Returns:
        Format string: 'csv', 'xlsx', 'json', 'parquet', or 'tsv'.

    Raises:
        FormatError: If the extension is not recognised.
    """
    ext = Path(filename).suffix.lower()
    fmt = _EXTENSION_MAP.get(ext)
    if fmt is None:
        raise FormatError(f"Unsupported file extension: '{ext}'. Supported: {list(_EXTENSION_MAP.keys())}")
    return fmt


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def load_csv(path: str | Path) -> pd.DataFrame:
    """Load a CSV file with encoding fallback (utf-8 -> latin-1 -> cp1252)."""
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise FormatError(f"Cannot decode CSV file: {path}")


def load_xlsx(path: str | Path, sheet_name: str | int | None = 0) -> pd.DataFrame:
    """Load an Excel worksheet using openpyxl.

    Args:
        path: Path to the .xlsx file.
        sheet_name: Sheet name or 0-based index.  Defaults to the first sheet.
    """
    try:
        return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    except Exception as e:
        raise FormatError(f"Failed to read XLSX file: {e}") from e


def load_json(path: str | Path) -> pd.DataFrame:
    """Load a JSON file with orientation auto-detection."""
    try:
        return pd.read_json(path)
    except ValueError:
        pass
    for orient in ("records", "columns", "index"):
        try:
            return pd.read_json(path, orient=orient)
        except (ValueError, KeyError, TypeError):
            continue
    raise FormatError(f"Cannot parse JSON file: {path}")


def load_parquet(path: str | Path) -> pd.DataFrame:
    """Load a Parquet file using pyarrow."""
    try:
        return pd.read_parquet(path, engine="pyarrow")
    except Exception as e:
        raise FormatError(f"Failed to read Parquet file: {e}") from e


def load_tsv(path: str | Path) -> pd.DataFrame:
    """Load a TSV (tab-separated values) file with encoding fallback."""
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return pd.read_csv(path, sep="\t", encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise FormatError(f"Cannot decode TSV file: {path}")


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

_LOADERS = {
    "csv": load_csv,
    "xlsx": load_xlsx,
    "json": load_json,
    "parquet": load_parquet,
    "tsv": load_tsv,
}


def ingest(file_path: str | Path, format: str | None = None) -> pd.DataFrame:
    """Load a file into a DataFrame, dispatching to the appropriate loader.

    Args:
        file_path: Path to the data file.
        format: Explicit format override.  When *None*, auto-detects from the
                file extension.

    Returns:
        DataFrame with the file contents.

    Raises:
        FormatError: If the format is unsupported or the file cannot be read.
    """
    path = Path(file_path)
    fmt = format or detect_format(path.name)

    loader = _LOADERS.get(fmt)
    if loader is None:
        raise FormatError(f"No loader registered for format: '{fmt}'")

    logger.info("Ingesting %s (format=%s)", path.name, fmt)
    return loader(path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def get_xlsx_sheet_names(path: str | Path) -> list[str]:
    """Return the sheet names in an Excel workbook."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        raise FormatError(f"Failed to read sheet names: {e}") from e


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------


def export(df: pd.DataFrame, format: str, output_path: str | Path) -> Path:
    """Serialise a DataFrame to the specified format.

    Args:
        df: Source DataFrame.
        format: Target format ('csv', 'xlsx', 'json', 'parquet', 'tsv').
        output_path: Destination file path.

    Returns:
        Path to the written file.

    Raises:
        FormatError: If the format is unsupported or writing fails.
    """
    path = Path(output_path)
    exporters: dict[str, callable] = {
        "csv": lambda: df.to_csv(path, index=False),
        "xlsx": lambda: df.to_excel(path, index=False, engine="openpyxl"),
        "json": lambda: df.to_json(path, orient="records", indent=2),
        "parquet": lambda: df.to_parquet(path, index=False, engine="pyarrow"),
        "tsv": lambda: df.to_csv(path, index=False, sep="\t"),
    }

    exporter = exporters.get(format)
    if exporter is None:
        raise FormatError(f"No exporter registered for format: '{format}'")

    try:
        exporter()
    except Exception as e:
        raise FormatError(f"Export to {format} failed: {e}") from e

    logger.info("Exported to %s (format=%s)", path.name, format)
    return path
